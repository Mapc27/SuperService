import json
import os
import time

import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from colorama import Fore, Style
from urllib3.exceptions import MaxRetryError, NewConnectionError

import config


def decor(func):
    def wrapper(*args, **kwargs):
        while True:
            try:
                return func(*args, **kwargs)
            except (ConnectionRefusedError, requests.exceptions.ProxyError, MaxRetryError,
                    NewConnectionError, TimeoutError) as error:
                print(Fore.RED + "Обрабатывается ошибка {}, не переживайте".format(error))

    return wrapper


@decor
def get_request(url):
    return json.loads(requests.get(url, headers=config.headers).text)


if __name__ == '__main__':
    fill = PatternFill(fill_type='solid', start_color='00ffff', end_color='00ffff')
    fill_header = PatternFill(fill_type='solid', start_color='c5d9f1', end_color='c5d9f1')
    border = Border(bottom=Side(border_style='thick', color='1775e6'))
    page = 1
    limit = 20
    timeout = 600
    if os.path.exists('table_recall.xlsx'):
        wb = openpyxl.load_workbook('table_recall.xlsx')
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['is_target', 'entrant_fullname', 'href', 'name_competitive_group', 'name_status', 'education_level',
                   'id'])
        for element in ws[1]:
            element.border = border
            element.fill = fill_header
    while True:
        was_print = False
        response = get_request(config.recall_url.format(page, limit))
        count_page = response['paginator']['count_page']
        for page in range(1, count_page + 1):
            response = get_request(config.recall_url.format(page, limit))
            for app in response['data']:
                need_append = True
                for row in ws:
                    if str(row[6].value) == str(app['id']):
                        need_append = False

                if need_append:
                    href = 'http://10.3.60.2/cabinets/university/application/{}'.format(app['id'])
                    is_target = False
                    if 'бюджет, цел,' in app['name_competitive_group']:
                        is_target = True
                    ws.append([is_target, app['entrant_fullname'], href, app['name_competitive_group'],
                               app['name_status'], app['education_level'], app['id']])
                    for element in ws[ws.max_row]:
                        element.fill = fill
                    print(Fore.GREEN + str(is_target), app['entrant_fullname'], href, app['name_competitive_group'],
                          app['name_status'], app['education_level'], app['id'])
                    was_print = True
        status = True
        counter = 1
        file_name = 'table_recall'
        new_name = file_name
        while status:
            try:
                wb.save('{}.xlsx'.format(new_name))
                status = False
            except PermissionError:
                new_name = file_name + '_' + str(counter)

        wb.save('table_recall.xlsx')
        if not was_print:
            print(Style.RESET_ALL)
            print("Новых отзывов согласий нет")
        for i in range(0, timeout + 1):
            style = Fore.GREEN
            if timeout - i < timeout // 3:
                style = Fore.RED
            elif timeout - i < timeout // 3 * 2:
                style = Fore.YELLOW
            print(style + '\rУ тебя есть {}s, чтобы посмотреть файл table_recall.xlsx'.format(timeout - i), end='', flush=True)
            time.sleep(1)
        print()
        print(Fore.LIGHTCYAN_EX + '=' * 100)
        print('=' * 100)
        print(Style.RESET_ALL)
