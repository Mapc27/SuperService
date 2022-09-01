import os
from abc import abstractmethod
from datetime import datetime
from typing import Optional, Any, List, Union

import numpy as np
import openpyxl
import requests
from pydantic import BaseModel

from config import PROXIES, HEADERS, ENTRANTS_LIST_URL, COOKIE_DICT, ENTRANT_DOCS_URL, ENTRANT_APPLICATIONS_URL, \
    ENTRANT_APPLICATION_DOCS_LIST_URL, ENTRANT_SET_ORIGINAL_URL, ENTRANT_SHORT_URL, ENTRANT_APPLICATION_EDIT_URL, \
    ENTRANT_APPLICATION_INFO_URL, APP_SEARCH_URL, APP_SEARCH_UID_URL, ENTRANT_APPLICATION_SET_STATUS_URL
from decorators import bad_connection, bad_response


class Doc(BaseModel):
    can_edit: bool
    checked: bool
    doc_number: Optional[str]
    doc_series: Optional[str]
    has_file: bool
    id: int
    id_check_status: Optional[int]
    id_document_type: int
    id_entrant: int
    id_organization: Optional[int]
    issue_date: Optional[str]
    mark: Optional[int]
    name_check_statuses: Optional[str]
    name_document_type: Optional[str]
    name_subject: Optional[str]
    name_table: Optional[str]
    uid_epgu: Optional[str]


class Application(BaseModel):
    agreed: bool
    agreed_date: Optional[str]
    app_number: str
    code_status: str
    comment: Optional[str]
    entrant_snils: Optional[str]
    id: int
    id_competitive_group: int
    id_status: int
    name_competitive_group: str
    name_organization_okso: str
    name_status: str
    rating: Any
    registration_date: str
    uid_epgu: int


class ApplicationDoc(BaseModel):
    checked: bool
    doc_number: Optional[str]
    doc_series: Optional[str]
    id: int
    id_check_status: Optional[int]
    id_document_type: int
    id_entrant: int
    issue_date: Optional[str]
    mark: Optional[int]
    name_check_statuses: Optional[str]
    name_document_type: str
    name_subject: Optional[str]


class GeneralMethods:
    def __init__(self, city: Union[str, int]):
        self.headers = HEADERS.copy()
        self.proxies = PROXIES.copy()

        self.headers['Cookie'] = COOKIE_DICT[city]

    @bad_response
    @bad_connection
    def get_request(self, url: str, headers=None):
        if not headers:
            headers = self.headers
        return requests.get(url, headers=headers, proxies=self.proxies)

    @bad_response
    @bad_connection
    def post_request(self, url: str, data: dict, headers=None):
        if not headers:
            headers = self.headers
        return requests.post(url, headers=headers, json=data, proxies=self.proxies)

    def get_entrant_id(self, snils: int) -> int:
        response = self.get_request(ENTRANTS_LIST_URL.format(snils)).json()
        return response['data'][0]['id']

    def get_entrant_orig_doc_status(self, entrant_id: int) -> int:
        response = self.get_request(ENTRANT_SHORT_URL.format(entrant_id)).json()
        return response['data']['orig_doc']

    def get_entrant_docs(self, entrant_id: int) -> List[Doc]:
        response_json = self.get_request(ENTRANT_DOCS_URL.format(entrant_id)).json()
        if len(response_json['data']) != 1:
            raise ResponseDataNotAloneException(f"Error with data: response_json['data'] = {response_json['data']}")
        entrant_docs = response_json['data'][0]['docs']

        return list(map(
            lambda entrant_doc: Doc.parse_obj(entrant_doc), entrant_docs
        ))

    def get_entrant_applications(self, entrant_id: int) -> List[Application]:
        response_json = self.get_request(ENTRANT_APPLICATIONS_URL.format(entrant_id)).json()
        entrant_applications = response_json['data']

        return list(map(
            lambda entrant_doc: Application.parse_obj(entrant_doc), entrant_applications
        ))

    def get_application_docs_list(self, application_id: int) -> List[ApplicationDoc]:
        response_json = self.get_request(ENTRANT_APPLICATION_DOCS_LIST_URL.format(application_id)).json()

        if len(response_json['data']['docs']) != 2:
            raise ResponseDataNotAloneException(
                f"Error with docs: response_json['data']['docs'] = {response_json['data']['docs']}"
            )

        application_docs_list = response_json['data']['docs'][1]['docs']
        return list(map(
            lambda entrant_doc: ApplicationDoc.parse_obj(entrant_doc), application_docs_list
        ))

    def set_originals(self, entrant_id, id_document: int, orig_doc_date: str):
        self.post_request(
            ENTRANT_SET_ORIGINAL_URL.format(entrant_id),
            data={"id_document": id_document, "orig_doc_date": orig_doc_date}
        )

    def set_agreed(self, app_id, agreed_date, agreed_time, agreed, headers=None):
        response = self.get_request(
            ENTRANT_APPLICATION_INFO_URL.format(app_id),
            headers=headers
        ).json()

        if not response['done']:
            raise SetAgreedException(response['message'])

        request_data = response['data']

        request_data['agreed'] = True if agreed else False
        request_data['agreed_date'] = agreed_date
        request_data['agreed_time'] = agreed_time
        request_data['id_application'] = request_data['id']

        response = self.post_request(
            ENTRANT_APPLICATION_EDIT_URL.format(app_id),
            data=request_data,
            headers=headers
        ).json()
        if not response['done']:
            raise SetAgreedException(response['message'])

    @staticmethod
    def str_to_datetime(agreed_date):
        for fmt in ["%d/%m/%Y %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"]:
            try:
                dt = datetime.strptime(str(agreed_date), fmt)
                return dt.isoformat() + '+03:00', dt.strftime("%H:%M")
            except ValueError:
                pass
        raise ValueError("Неверный формат даты")

    def get_app_from_uid(self, app_uid, headers=None):
        try:
            apps = self.get_request(
                url=APP_SEARCH_URL.format(app_uid),
                headers=headers
            ).json()['data']
        except IndexError:
            try:
                apps = self.get_request(
                    url=APP_SEARCH_UID_URL.format(app_uid),
                    headers=headers
                ).json()['data']
            except IndexError:
                return

        app_list = []
        for app in apps:
            if str(app_uid) in [str(app['app_number']), str(app['uid_epgu'])]:
                app_list.append(app)

        if len(app_list) > 1:
            raise TooManyApplicationsFoundException("len(app_list) =", len(app_list))

        return app_list[0]

    def set_status_in_competition(self, application_id, headers=None):
        status = "in_competition"
        notification = {
            "id_template": None,
            "comment": "Публикация конкурсных списков ожидается: ",
            "id_notices_types": 10
        }
        status_comment = None
        return self.post_request(
            ENTRANT_APPLICATION_SET_STATUS_URL.format(application_id),
            data={"code": status, "notification": notification, "status_comment": status_comment},
            headers=headers
        )

    def set_status_out_competition(self, application_id, headers=None):
        status = "out_competition"
        notification = {
            "id_template": 41,
            "comment": "Поступающий не прошёл по конкурсу",
            "id_notices_types": 9
        }
        status_comment = None
        return self.post_request(
            ENTRANT_APPLICATION_SET_STATUS_URL.format(application_id),
            data={"code": status, "notification": notification, "status_comment": status_comment},
            headers=headers
        )

    def set_status_service_denied(self, application_id, headers=None):
        status = "service_denied"
        notification = {
            "id_template": 41,
            "comment": "Поступающий не прошёл по конкурсу",
            "id_notices_types": 9
        }
        status_comment = None
        return self.post_request(
            ENTRANT_APPLICATION_SET_STATUS_URL.format(application_id),
            data={"code": status, "notification": notification, "status_comment": status_comment},
            headers=headers
        )


class SuperService:
    def __init__(self, city_name, process_number, result_filename=None):
        self.general_methods = GeneralMethods(city_name)
        self.city_name = city_name
        self.process_number = process_number
        self.result_filename = result_filename

    @staticmethod
    def get_data_from_file(file_name, start_letter, end_letter) -> list:
        excel_file = openpyxl.load_workbook(file_name)
        sheet = excel_file.active
        array = list(sheet[f"{start_letter + str(sheet.min_row)}":f"{end_letter + str(sheet.max_row)}"])
        return excel_file, array

    def divide_file(self):
        excel_file = openpyxl.load_workbook(f"{self.city_name}.xlsx")
        sheet = excel_file.active

        splits = np.array_split(list(sheet.rows), self.process_number)

        file_names = []
        file_number = 0
        for split in splits:
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in split:
                ws.append([cell.value for cell in row])

            file_name = f"{self.city_name}_{file_number}.xlsx"
            wb.save(file_name)
            file_names.append(file_name)
            file_number += 1
        return file_names

    def fold_files(self, file_names):
        wb = openpyxl.Workbook()
        ws = wb.active
        for file_name in file_names:
            excel_file = openpyxl.load_workbook(file_name)
            sheet = excel_file.active
            for row in sheet:
                ws.append([cell.value for cell in row])
            os.remove(file_name)
        result_filename = self.result_filename + ".xlsx" if self.result_filename else f"{self.city_name}_result.xlsx"
        wb.save(result_filename)

    @abstractmethod
    def main(self):
        pass


class HasUncheckedExamResultsException(BaseException):
    pass


class HasNotCheckedCertificateException(BaseException):
    pass


class ResponseDataNotAloneException(BaseException):
    pass


class SetAgreedException(BaseException):
    pass


class TooManyApplicationsFoundException(BaseException):
    pass
