from multiprocessing import Process
from pprint import pprint
from typing import List

import numpy as np
import openpyxl

from config import CERTIFICATE_DOCUMENT_TYPE_ID
from general import GeneralMethods, Doc, HasNotCheckedCertificateException, ResponseDataNotAloneException


class SuperServiceOriginal:
    def __init__(self, institute_type, city_name, process_number, document_type_id):
        self.general_methods = GeneralMethods(institute_type)
        self.city_name = city_name
        self.process_number = process_number
        self.document_type_id = document_type_id

    @staticmethod
    def get_data_from_file(file_name) -> list:
        excel_file = openpyxl.load_workbook(file_name)
        sheet = excel_file.active
        array = list(sheet[f"A{sheet.min_row + 1}":f"G{sheet.max_row}"])
        return excel_file, array

    def divide_file(self):
        excel_file = openpyxl.load_workbook(f"{self.city_name}.xlsx")
        sheet = excel_file.active

        splits = np.array_split(list(sheet.rows), self.process_number)

        file_names = []
        file_number = 0
        for split in splits:
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in split:
                ws.append([cell.value for cell in row])

            file_name = f"{self.city_name}_{file_number}.xlsx"
            wb.save(file_name)
            file_names.append(file_name)
            file_number += 1
        return file_names

    def main(self):
        file_names = self.divide_file()
        process_list = []

        for file_name in file_names:
            process = Process(target=self.entrant_loop, args=(file_name,))
            process_list.append(process)

        for process in process_list:
            process.start()

        for process in process_list:
            process.join()

    def entrant_loop(self, file_name):
        excel_file, rows = self.get_data_from_file(file_name)
        len_rows = len(rows)
        print("len_rows", len_rows)

        for i in range(len_rows):
            print(f"{i} of {len_rows}")
            try:
                self.process_entrant(rows[i])
            except Exception as e:
                print(e)
                rows[i][6].value = "Exception"
                pass

        print(f"Saving file {file_name}")
        excel_file.save(file_name)

    @staticmethod
    def validate_certificate_series(certificate_series):
        if not certificate_series:
            return ''

        certificate_series = certificate_series.replace("-", "")
        return certificate_series

    def process_entrant(self, row):
        snils, orig_doc_date, certificate_series, certificate_number, empty_cell = \
            row[1].value, row[2].value, row[3].value, row[4].value, row[6]
        orig_doc_date = orig_doc_date.replace('.', ':')
        certificate_series = self.validate_certificate_series(certificate_series)
        print("snils:", snils)

        try:
            entrant_id = self.general_methods.get_entrant_id(snils)
            print("entrant_id =", entrant_id)
        except IndexError:
            print("entrant_id not found")
            empty_cell.value = "entrant_id not found"
            return

        orig_doc_status = self.general_methods.get_entrant_orig_doc_status(entrant_id)
        if orig_doc_status:
            print("original doc status already is True")
            return
        else:
            print("orig_doc_status =", orig_doc_status)

        try:
            entrant_docs = self.general_methods.get_entrant_docs(entrant_id)
            print("documents found")
        except ResponseDataNotAloneException:
            print("documents not found")
            empty_cell.value = "documents not found"
            return

        checked_certificates = self.get_checked_certificates(entrant_docs)

        if not checked_certificates:
            print("Has Not Checked Certificate")
            certificates = self.get_certificates_by_data(entrant_docs, certificate_series, certificate_number)
            if not certificates:
                print("Can not find certificate")
                empty_cell.value = "Can not find certificate"
                return
            certificate = certificates[0]
            print("Certificate with data found")
        else:
            certificate = checked_certificates[0]
            print('checked certificate found')

        print("Set_originals")
        self.general_methods.set_originals(entrant_id, certificate.id, orig_doc_date)

    def get_checked_certificates(self, entrant_docs: List[Doc]):
        return list(filter(
            lambda entrant_doc: entrant_doc.id_document_type == self.document_type_id and entrant_doc.checked,
            entrant_docs
        ))

    def get_certificates_by_data(self, entrant_docs: List[Doc], doc_series, doc_number):
        def concatenate(series, number):
            if not series:
                series = ''
            if not number:
                number = ''
            return str(series) + str(number)

        return list(filter(
            lambda entrant_doc:
            concatenate(entrant_doc.doc_series, entrant_doc.doc_number) == doc_series + doc_number and
            entrant_doc.id_document_type == self.document_type_id,
            entrant_docs
        ))

    @staticmethod
    def has_unchecked_certificate(checked_certificates: List[Doc]):
        if len(checked_certificates) == 0:
            raise HasNotCheckedCertificateException("Entrant hasn't checked certificates")


if __name__ == '__main__':
    SuperServiceOriginal(institute_type=1, city_name='chelny', process_number=1, document_type_id=10).main()
