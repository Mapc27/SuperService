import time
from multiprocessing import Process

import numpy as np
import openpyxl

from general import GeneralMethods

host_url = 'http://10.3.60.2/'
# host_url = 'http://85.142.162.4:8032/'

entrants_list_url = host_url + 'api/entrants/list?page={}&limit=20'
entrant_applications_url = host_url + 'api/entrants/{}/applications'
entrant_application_set_status_url = host_url + 'api/applications/{}/status/set'

entrant_url = host_url + "api/entrants/list?page=1&limit=20&search_snils={}"
app_search_url = host_url + "api/applications/list?page=1&limit=20&search_snils=&search_number={}"
app_search_uid_url = host_url + "api/applications/list?page=1&limit=20&search_uid_epgu={}"


general_methods = GeneralMethods(institute_type=0)


def get_app_from_uid(app_uid):
    try:
        apps = general_methods.get_request(app_search_url.format(app_uid)).json()['data']
    except IndexError:
        try:
            apps = general_methods.get_request(app_search_uid_url.format(app_uid)).json()['data']
        except IndexError:
            return

    for app in apps:
        if str(app_uid) in [str(app['app_number']), str(app['uid_epgu'])]:
            return app


def set_status_in_competition(application_id):
    status = "in_competition"
    notification = {
        "id_template": None,
        "comment": "Публикация конкурсных списков ожидается: ",
        "id_notices_types": 10
    }
    status_comment = None
    return general_methods.post_request(
        entrant_application_set_status_url.format(application_id),
        data={"code": status, "notification": notification, "status_comment": status_comment},
    )

# def set_status_to_app_call_off(application_id):
#     status = "app_call_off"
#     return general_methods.post_request(
#         entrant_application_set_status_url.format(application_id),
#         data={"code": status},
#     )


def process_application(app_uid, result_cell):
    app = get_app_from_uid(app_uid)
    if not app:
        return
    # considered
    if app['id_status'] == 4:
        set_status_in_competition(app['id'])
        app = get_app_from_uid(app_uid)
        print('Set status')

    result_cell.value = app['id_status']


def loop(rows):
    for row in rows:
        app_uid = row[0].value.replace('appid', '')
        result_cell = row[1]
        print('app_uid:', app_uid)
        try:
            if app_uid:
                process_application(app_uid, result_cell)
            else:
                print("Cannot get app_uid")
        except Exception as exception:
            print("Возникла ошибка при обработке app_uid: ", app_uid, exception)
            print("Продолжаю дальше ...")
    print("Process ended")


def map_file(city, process_number):
    excel_file = openpyxl.load_workbook(f"{city}.xlsx")
    sheet = excel_file.active

    splits = np.array_split(list(sheet.rows), process_number)

    file_number = 0
    for split in splits:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in split:
            ws.append((row[0].value, ))
        wb.save(f"{city}_{file_number}.xlsx")
        file_number += 1


def main(file_name):
    excel_file = openpyxl.load_workbook(file_name)

    sheet = excel_file.active

    array = list(sheet[f"A{sheet.min_row + 1}":f"B{sheet.max_row}"])

    print(len(array))

    try:
        loop(array)
    except Exception as exception:
        raise exception
    finally:
        print("Saving file")
        excel_file.save(file_name)

    print("Main ended successfully, program will wait 30s and then close")
    time.sleep(30)


if __name__ == '__main__':
    city_name = 'kazan'
    process_number_ = 10
    map_file(city_name, process_number_)
    process_list = []

    for i in range(process_number_):
        process = Process(target=main, args=(f"{city_name}_{i}.xlsx", ))
        process_list.append(process)

    for process in process_list:
        process.start()

    for process in process_list:
        process.join()
